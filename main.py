"""
Greenwashing Auditor - FastAPI web app

The user brings their own Anthropic API key (never stored). The backend crawls
the target site, streams live progress via Server-Sent Events, and produces an
Excel report evaluating environmental claims under Australian law.
"""

from __future__ import annotations

import asyncio
import csv
import io
import json
import os
import re
import secrets
import time
import uuid
import xml.etree.ElementTree as ET
from collections import deque
from dataclasses import dataclass, field
from datetime import datetime, timedelta
from typing import AsyncIterator, Iterable
from urllib.parse import urljoin, urlparse, urldefrag
from urllib.robotparser import RobotFileParser

import anthropic
import httpx
from bs4 import BeautifulSoup
from fastapi import FastAPI, HTTPException, Request
from fastapi.exceptions import RequestValidationError
from fastapi.responses import HTMLResponse, StreamingResponse, Response, JSONResponse
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from pydantic import BaseModel, Field


# --------------------------------------------------------------------------- #
# Config
# --------------------------------------------------------------------------- #

MODEL = "claude-sonnet-4-5"
USER_AGENT = "GreenwashingAuditor/1.0 (+compliance review)"
REQUEST_TIMEOUT = 20
MAX_PAGE_CHARS = 20_000
HARD_MAX_PAGES = 300  # absolute ceiling regardless of user input
JOB_TTL_SECONDS = 60 * 60  # keep finished job results for 1 hour


SKIP_EXTENSIONS = (
    ".pdf", ".jpg", ".jpeg", ".png", ".gif", ".webp", ".svg",
    ".zip", ".mp3", ".mp4", ".mov", ".avi", ".css", ".js",
    ".ico", ".xml", ".woff", ".woff2", ".ttf",
)


SYSTEM_PROMPT = """You are a senior Australian consumer law compliance analyst \
specialising in greenwashing. You review web page content and identify \
environmental or sustainability claims that may breach Australian law or \
regulator guidance.

Legal framework you apply:
  - Australian Consumer Law, sections 18 and 29 of Schedule 2 to the \
Competition and Consumer Act 2010 (Cth): prohibits misleading or deceptive \
conduct and false or misleading representations.
  - ACCC "Making environmental claims: A guide for business" (December 2023). \
The 8 principles:
      1. Make accurate and truthful claims
      2. Have evidence to back up your claims
      3. Don't hide or omit important information
      4. Explain any conditions or qualifications on your claims
      5. Avoid broad and unqualified claims (e.g. "green", "eco-friendly", \
"sustainable", "carbon neutral" without substantiation)
      6. Use clear and easy-to-understand language
      7. Visual elements should not create a misleading impression
      8. Be direct and open about your sustainability transition (no \
aspirational statements presented as current fact)
  - ASIC Information Sheet 271 (INFO 271) applies where the page relates to \
financial products or investments described as sustainable, ethical, ESG, \
green or similar.

How to flag a claim:
  - Quote the claim VERBATIM (short, under 15 words; trim with ellipsis if longer).
  - Identify the specific principle(s) or statutory provision likely engaged.
  - Explain concisely why the claim is at risk.
  - Assign severity: "High", "Medium", or "Low".
  - Suggest a concrete remediation.

If the page contains NO environmental/sustainability claims, return an empty \
list. Do not invent claims. Do not flag clearly substantiated claims (e.g. \
those citing a recognised standard, certification, or disclosed methodology) \
unless there is a specific problem with the substantiation.

Respond with a single JSON object only, no prose before or after:

{
  "claims": [
    {
      "claim_text": "string - verbatim quote",
      "breach_reason": "string - which principle(s) and why",
      "severity": "High" | "Medium" | "Low",
      "remediation": "string - concrete fix"
    }
  ]
}
"""


USER_PROMPT_TEMPLATE = """Review the following web page for potential \
greenwashing under Australian law. Return JSON only.

URL: {url}
Page title: {title}

Page text content:
---
{content}
---
"""


# --------------------------------------------------------------------------- #
# Models
# --------------------------------------------------------------------------- #

class StartAuditRequest(BaseModel):
    url: str
    api_key: str = Field(min_length=10)
    max_pages: int = Field(default=50, ge=1, le=HARD_MAX_PAGES)
    delay: float = Field(default=0.5, ge=0.0, le=10.0)
    include_subdomains: bool = False
    previous_upload_id: str | None = None  # ID from /api/upload_previous


@dataclass
class PageContent:
    url: str
    title: str
    text: str


@dataclass
class FlaggedClaim:
    url: str
    page_title: str
    claim_text: str
    breach_reason: str
    severity: str
    remediation: str
    status: str = "NEW"  # NEW | KNOWN | RESOLVED


@dataclass
class PreviousClaim:
    """A claim loaded from a previously-uploaded Excel report."""
    url_key: str           # normalised URL for matching
    token_set: frozenset   # normalised word tokens of the claim text
    severity: str
    claim_text: str        # original, for carry-through if marked RESOLVED
    page_title: str
    breach_reason: str
    remediation: str
    matched: bool = False  # set True if a new finding matches this one


@dataclass
class Job:
    id: str
    request: StartAuditRequest
    queue: asyncio.Queue = field(default_factory=asyncio.Queue)
    # Replay buffer: every event is also appended here so a reconnecting
    # client can catch up by passing ?since=<N> in the stream URL.
    event_log: list[dict] = field(default_factory=list)
    flags: list[FlaggedClaim] = field(default_factory=list)
    previous_claims: list[PreviousClaim] = field(default_factory=list)
    pages_scanned: int = 0
    status: str = "pending"  # pending | running | done | error
    error_message: str | None = None
    excel_bytes: bytes | None = None
    csv_bytes: bytes | None = None
    created_at: datetime = field(default_factory=datetime.utcnow)
    finished_at: datetime | None = None


# --------------------------------------------------------------------------- #
# In-memory job store (sufficient for single-instance Railway deployment)
# --------------------------------------------------------------------------- #

JOBS: dict[str, Job] = {}


def _sweep_old_jobs() -> None:
    cutoff = datetime.utcnow() - timedelta(seconds=JOB_TTL_SECONDS)
    for jid, job in list(JOBS.items()):
        if job.finished_at and job.finished_at < cutoff:
            JOBS.pop(jid, None)


# --------------------------------------------------------------------------- #
# Diff / known-claim matching
# --------------------------------------------------------------------------- #

MATCH_THRESHOLD = 0.5  # Jaccard similarity above this = same claim

# Words to ignore when fingerprinting (common English stop words + filler)
_STOP_WORDS = frozenset({
    "the", "a", "an", "and", "or", "but", "is", "are", "was", "were",
    "be", "been", "being", "have", "has", "had", "do", "does", "did",
    "will", "would", "could", "should", "may", "might", "of", "in", "on",
    "at", "to", "for", "with", "by", "from", "as", "that", "this",
    "these", "those", "it", "its", "our", "we", "you", "your",
})


def _normalise_url_for_match(url: str) -> str:
    """Aggressively normalise a URL for matching: lowercase, strip scheme,
    strip www, strip trailing slash and fragment, strip query string."""
    if not url:
        return ""
    try:
        p = urlparse(url.strip())
        netloc = p.netloc.lower()
        if netloc.startswith("www."):
            netloc = netloc[4:]
        path = p.path.rstrip("/").lower()
        return f"{netloc}{path}"
    except Exception:
        return url.strip().lower()


def _tokenise_claim(text: str) -> frozenset:
    """Turn a claim string into a set of meaningful tokens for Jaccard matching.
    Hyphens are treated as spaces so 'eco-friendly' == 'eco friendly'."""
    if not text:
        return frozenset()
    t = text.strip().strip('"\u201c\u201d\'').lower()
    t = re.sub(r"\.\.\.$", "", t)
    # Treat hyphens and underscores as word separators
    t = t.replace("-", " ").replace("_", " ")
    # Keep letters/numbers only
    t = re.sub(r"[^\w\s]", " ", t)
    tokens = {w for w in t.split() if len(w) > 1 and w not in _STOP_WORDS}
    return frozenset(tokens)


def _jaccard(a: frozenset, b: frozenset) -> float:
    if not a or not b:
        return 0.0
    inter = len(a & b)
    union = len(a | b)
    return inter / union if union else 0.0


def _classify_new_flag(flag: FlaggedClaim, previous: list[PreviousClaim]) -> str:
    """Return 'KNOWN' if flag matches a previous claim on same URL, else 'NEW'.
    Also marks the matched previous claim so we can detect RESOLVED ones later."""
    if not previous:
        return "NEW"
    flag_url = _normalise_url_for_match(flag.url)
    flag_tokens = _tokenise_claim(flag.claim_text)
    best_score = 0.0
    best_match = None
    for prev in previous:
        if prev.matched:
            continue
        if prev.url_key != flag_url:
            continue
        score = _jaccard(flag_tokens, prev.token_set)
        if score > best_score:
            best_score = score
            best_match = prev
    if best_match is not None and best_score >= MATCH_THRESHOLD:
        best_match.matched = True
        return "KNOWN"
    return "NEW"


def parse_previous_excel(data: bytes) -> list[PreviousClaim]:
    """Parse an Excel report previously produced by this tool.
    Expects a 'Findings' sheet with columns:
      URL | Page Title | Severity | Claim (verbatim) | Why it may breach | Suggested remediation
    (An optional 'Status' column is ignored on load.)
    """
    from openpyxl import load_workbook
    try:
        wb = load_workbook(io.BytesIO(data), data_only=True, read_only=True)
    except Exception as e:
        raise ValueError(f"Could not open Excel file: {e}")

    # Prefer 'Findings' sheet, fall back to first sheet
    sheet_name = "Findings" if "Findings" in wb.sheetnames else wb.sheetnames[0]
    ws = wb[sheet_name]

    rows = ws.iter_rows(values_only=True)
    try:
        header = next(rows)
    except StopIteration:
        return []
    if not header:
        return []

    # Map common header names to indices (case-insensitive)
    header_map = {}
    for i, cell in enumerate(header):
        if cell is None:
            continue
        key = str(cell).strip().lower()
        header_map[key] = i

    def pick(*names: str) -> int | None:
        for n in names:
            if n in header_map:
                return header_map[n]
        # substring fallback
        for n in names:
            for k, v in header_map.items():
                if n in k:
                    return v
        return None

    i_url = pick("url")
    i_title = pick("page title", "title")
    i_sev = pick("severity")
    i_claim = pick("claim (verbatim)", "claim verbatim", "claim")
    i_reason = pick("why it may breach", "breach reason", "reason")
    i_remed = pick("suggested remediation", "remediation", "fix")

    if i_url is None or i_claim is None:
        raise ValueError(
            "Could not find required columns in Excel. Expected a 'Findings' "
            "sheet with at least 'URL' and 'Claim (verbatim)' columns."
        )

    results: list[PreviousClaim] = []
    for row in rows:
        if row is None:
            continue
        try:
            url_val = row[i_url]
        except IndexError:
            continue
        if not url_val:
            continue
        claim_val = row[i_claim] if i_claim < len(row) else None
        if not claim_val:
            continue
        results.append(PreviousClaim(
            url_key=_normalise_url_for_match(str(url_val)),
            token_set=_tokenise_claim(str(claim_val)),
            severity=str(row[i_sev] or "").strip() if i_sev is not None and i_sev < len(row) else "",
            claim_text=str(claim_val).strip(),
            page_title=str(row[i_title] or "").strip() if i_title is not None and i_title < len(row) else "",
            breach_reason=str(row[i_reason] or "").strip() if i_reason is not None and i_reason < len(row) else "",
            remediation=str(row[i_remed] or "").strip() if i_remed is not None and i_remed < len(row) else "",
        ))
    return results


# --------------------------------------------------------------------------- #
# Crawling helpers
# --------------------------------------------------------------------------- #

def _normalise_url(url: str) -> str:
    url, _ = urldefrag(url)
    return url.rstrip("/")


def _is_same_site(candidate: str, start_netloc: str, include_subdomains: bool) -> bool:
    cand_netloc = urlparse(candidate).netloc.lower()
    start_netloc = start_netloc.lower()
    if cand_netloc == start_netloc:
        return True
    if include_subdomains and cand_netloc.endswith("." + start_netloc):
        return True
    return False


def _looks_like_html(url: str) -> bool:
    path = urlparse(url).path.lower()
    return not path.endswith(SKIP_EXTENSIONS)


async def _fetch(client: httpx.AsyncClient, url: str) -> httpx.Response | None:
    try:
        return await client.get(url, timeout=REQUEST_TIMEOUT, follow_redirects=True)
    except (httpx.HTTPError, httpx.InvalidURL):
        return None


async def discover_sitemap_urls(client: httpx.AsyncClient, base_url: str) -> list[str]:
    found: set[str] = set()
    candidates: list[str] = []

    parsed = urlparse(base_url)
    robots_url = f"{parsed.scheme}://{parsed.netloc}/robots.txt"
    r = await _fetch(client, robots_url)
    if r is not None and r.status_code == 200:
        for line in r.text.splitlines():
            if line.lower().startswith("sitemap:"):
                candidates.append(line.split(":", 1)[1].strip())
    candidates.append(f"{parsed.scheme}://{parsed.netloc}/sitemap.xml")

    seen: set[str] = set()
    queue = deque(candidates)
    while queue:
        sm_url = queue.popleft()
        if sm_url in seen:
            continue
        seen.add(sm_url)
        r = await _fetch(client, sm_url)
        if r is None or r.status_code != 200:
            continue
        try:
            root = ET.fromstring(r.content)
        except ET.ParseError:
            continue
        tag = root.tag.lower().split("}", 1)[-1]
        if tag == "sitemapindex":
            for loc in root.iter():
                if loc.tag.lower().endswith("loc") and loc.text:
                    queue.append(loc.text.strip())
        elif tag == "urlset":
            for loc in root.iter():
                if loc.tag.lower().endswith("loc") and loc.text:
                    found.add(_normalise_url(loc.text.strip()))
    return sorted(found)


def _extract_text(html: str) -> tuple[str, str]:
    soup = BeautifulSoup(html, "lxml")
    for tag in soup(["script", "style", "noscript"]):
        tag.decompose()
    title = (soup.title.string.strip() if soup.title and soup.title.string else "")

    alt_texts = []
    for img in soup.find_all("img"):
        alt = (img.get("alt") or "").strip()
        if alt:
            alt_texts.append(f"[image alt: {alt}]")

    body_text = soup.get_text(separator=" ", strip=True)
    body_text = re.sub(r"\s+", " ", body_text)
    combined = body_text
    if alt_texts:
        combined += "\n\n" + "\n".join(alt_texts)
    if len(combined) > MAX_PAGE_CHARS:
        combined = combined[:MAX_PAGE_CHARS] + " [...truncated]"
    return title, combined


def _extract_links(html: str, base_url: str) -> Iterable[str]:
    soup = BeautifulSoup(html, "lxml")
    for a in soup.find_all("a", href=True):
        href = a["href"].strip()
        if not href or href.startswith(("mailto:", "tel:", "javascript:", "#")):
            continue
        yield _normalise_url(urljoin(base_url, href))


# --------------------------------------------------------------------------- #
# Claude analysis
# --------------------------------------------------------------------------- #

def analyse_page_sync(
    client: anthropic.Anthropic, page: PageContent
) -> list[FlaggedClaim]:
    """Synchronous Anthropic call (SDK is sync). We run this in a thread."""
    user_msg = USER_PROMPT_TEMPLATE.format(
        url=page.url,
        title=page.title or "(no title)",
        content=page.text or "(no text)",
    )
    try:
        resp = client.messages.create(
            model=MODEL,
            max_tokens=2048,
            system=SYSTEM_PROMPT,
            messages=[{"role": "user", "content": user_msg}],
        )
    except anthropic.APIError:
        return []

    raw = "".join(b.text for b in resp.content if b.type == "text").strip()
    if raw.startswith("```"):
        raw = re.sub(r"^```(?:json)?\s*|\s*```$", "", raw, flags=re.MULTILINE).strip()
    try:
        parsed = json.loads(raw)
    except json.JSONDecodeError:
        return []

    out: list[FlaggedClaim] = []
    for c in parsed.get("claims", []) if isinstance(parsed, dict) else []:
        if not isinstance(c, dict):
            continue
        out.append(FlaggedClaim(
            url=page.url,
            page_title=page.title,
            claim_text=str(c.get("claim_text", "")).strip(),
            breach_reason=str(c.get("breach_reason", "")).strip(),
            severity=str(c.get("severity", "")).strip() or "Medium",
            remediation=str(c.get("remediation", "")).strip(),
        ))
    return out


# --------------------------------------------------------------------------- #
# Excel writer
# --------------------------------------------------------------------------- #

SEVERITY_FILL = {
    "High": PatternFill("solid", start_color="F8CBAD"),
    "Medium": PatternFill("solid", start_color="FFE699"),
    "Low": PatternFill("solid", start_color="C6E0B4"),
}

STATUS_FILL = {
    "NEW":      PatternFill("solid", start_color="FCE4EC"),  # pink-ish
    "KNOWN":    PatternFill("solid", start_color="E7E6E6"),  # muted grey
    "RESOLVED": PatternFill("solid", start_color="E2EFDA"),  # pale green
}


def build_csv(flags: list[FlaggedClaim], start_url: str) -> bytes:
    buf = io.StringIO()
    writer = csv.writer(buf, quoting=csv.QUOTE_ALL)
    writer.writerow([
        "Status", "URL", "Page Title", "Severity", "Claim",
        "Breach reason (ACCC / ACL / ASIC)", "Suggested remediation",
    ])
    sev_order = {"High": 0, "Medium": 1, "Low": 2}
    status_order = {"NEW": 0, "KNOWN": 1, "RESOLVED": 2}
    for f in sorted(flags, key=lambda x: (status_order.get(x.status, 3),
                                          sev_order.get(x.severity, 3),
                                          x.url)):
        writer.writerow([
            f.status, f.url, f.page_title, f.severity, f.claim_text,
            f.breach_reason, f.remediation,
        ])
    return buf.getvalue().encode("utf-8-sig")


def build_excel(flags: list[FlaggedClaim], pages_scanned: int, start_url: str) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Findings"

    headers = [
        "Status", "URL", "Page Title", "Severity", "Claim (verbatim)",
        "Why it may breach (ACCC / ACL / ASIC)", "Suggested remediation",
    ]
    ws.append(headers)
    header_font = Font(name="Arial", bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", start_color="305496")
    for col_idx in range(1, len(headers) + 1):
        c = ws.cell(row=1, column=col_idx)
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(horizontal="left", vertical="center")

    sev_order = {"High": 0, "Medium": 1, "Low": 2}
    status_order = {"NEW": 0, "KNOWN": 1, "RESOLVED": 2}
    sorted_flags = sorted(
        flags,
        key=lambda x: (status_order.get(x.status, 3),
                       sev_order.get(x.severity, 3),
                       x.url),
    )

    for f in sorted_flags:
        ws.append([f.status, f.url, f.page_title, f.severity,
                   f.claim_text, f.breach_reason, f.remediation])
        r = ws.max_row

        # Status cell formatting
        status_cell = ws.cell(row=r, column=1)
        status_fill = STATUS_FILL.get(f.status)
        if status_fill is not None:
            status_cell.fill = status_fill
            status_cell.font = Font(name="Arial", bold=True)

        # URL hyperlink (skip for resolved ones without a real URL)
        url_cell = ws.cell(row=r, column=2)
        if f.url and f.url.startswith(("http://", "https://")):
            url_cell.hyperlink = f.url
            url_cell.font = Font(name="Arial", color="0563C1", underline="single")

        # Severity cell colouring
        sev_cell = ws.cell(row=r, column=4)
        fill = SEVERITY_FILL.get(f.severity)
        if fill is not None:
            sev_cell.fill = fill
            sev_cell.font = Font(name="Arial", bold=True)

        # Strike-through for resolved rows
        if f.status == "RESOLVED":
            for col in range(1, len(headers) + 1):
                cell = ws.cell(row=r, column=col)
                cell.font = Font(name="Arial", strike=True, color="808080")

        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=r, column=col)
            if cell.font.name != "Arial":
                cell.font = Font(name="Arial")
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    widths = [12, 45, 30, 10, 55, 60, 55]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    summary = wb.create_sheet("Summary")
    summary["A1"] = "Greenwashing Audit Summary"
    summary["A1"].font = Font(name="Arial", bold=True, size=14)
    summary["A3"], summary["B3"] = "Target site", start_url
    summary["A4"], summary["B4"] = "Pages scanned", pages_scanned
    summary["A5"], summary["B5"] = "Total findings", len(flags)
    summary["A6"], summary["B6"] = "NEW (since last audit)", f'=COUNTIF(Findings!A:A,"NEW")'
    summary["A7"], summary["B7"] = "KNOWN (matched previous)", f'=COUNTIF(Findings!A:A,"KNOWN")'
    summary["A8"], summary["B8"] = "RESOLVED (gone since last audit)", f'=COUNTIF(Findings!A:A,"RESOLVED")'
    summary["A10"], summary["B10"] = "High severity", f'=COUNTIF(Findings!D:D,"High")'
    summary["A11"], summary["B11"] = "Medium severity", f'=COUNTIF(Findings!D:D,"Medium")'
    summary["A12"], summary["B12"] = "Low severity", f'=COUNTIF(Findings!D:D,"Low")'
    summary["A14"] = "Framework applied"
    summary["A15"] = "Australian Consumer Law ss 18 & 29 (Sch 2, CCA 2010)"
    summary["A16"] = "ACCC 'Making environmental claims' (Dec 2023) - 8 principles"
    summary["A17"] = "ASIC Information Sheet 271 (financial products)"
    for row in range(3, 18):
        summary.cell(row=row, column=1).font = Font(name="Arial")
    summary.column_dimensions["A"].width = 45
    summary.column_dimensions["B"].width = 60

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# --------------------------------------------------------------------------- #
# Main audit coroutine
# --------------------------------------------------------------------------- #

async def run_audit(job: Job) -> None:
    req = job.request
    q = job.queue

    async def emit(event: str, data: dict) -> None:
        evt = {"event": event, "data": data}
        job.event_log.append(evt)
        await q.put(evt)

    job.status = "running"
    await emit("status", {"message": f"Starting audit of {req.url}"})

    parsed_start = urlparse(req.url)
    if not parsed_start.scheme or not parsed_start.netloc:
        job.status = "error"
        job.error_message = "Invalid URL. Include scheme (https://)."
        job.finished_at = datetime.utcnow()
        await emit("error", {"message": job.error_message})
        await emit("done", {})
        return

    start_netloc = parsed_start.netloc

    # robots.txt
    rp = RobotFileParser()
    rp.set_url(f"{parsed_start.scheme}://{start_netloc}/robots.txt")
    try:
        await asyncio.to_thread(rp.read)
    except Exception:
        pass

    def allowed(url: str) -> bool:
        try:
            return rp.can_fetch(USER_AGENT, url)
        except Exception:
            return True

    # Validate API key early with a tiny ping
    anth = anthropic.Anthropic(api_key=req.api_key)

    async with httpx.AsyncClient(
        headers={"User-Agent": USER_AGENT},
        follow_redirects=True,
    ) as http:
        await emit("status", {"message": "Discovering URLs from sitemap(s)..."})
        sitemap_urls = await discover_sitemap_urls(http, req.url)
        await emit("status", {
            "message": f"Found {len(sitemap_urls)} URL(s) in sitemaps"
        })

        start = _normalise_url(req.url)
        seen: set[str] = {start}
        queue: deque[str] = deque([start])

        for u in sitemap_urls:
            if u in seen:
                continue
            if not _is_same_site(u, start_netloc, req.include_subdomains):
                continue
            if not _looks_like_html(u):
                continue
            seen.add(u)
            queue.append(u)

        await emit("status", {"message": f"Queue seeded with {len(queue)} URL(s). Beginning crawl..."})
        await emit("total", {"planned": min(len(queue), req.max_pages)})

        pages_analysed = 0

        while queue and pages_analysed < req.max_pages:
            url = queue.popleft()
            if not allowed(url):
                await emit("skip", {"url": url, "reason": "robots.txt"})
                continue
            if not _looks_like_html(url):
                continue

            await emit("fetch", {"url": url, "index": pages_analysed + 1})
            r = await _fetch(http, url)
            if r is None or r.status_code != 200:
                await emit("skip", {
                    "url": url,
                    "reason": f"HTTP {r.status_code if r else 'error'}"
                })
                continue
            if "html" not in r.headers.get("Content-Type", "").lower():
                continue

            title, text = _extract_text(r.text)
            page = PageContent(url=url, title=title, text=text)

            await emit("analyse", {"url": url, "title": title})
            flags = await asyncio.to_thread(analyse_page_sync, anth, page)

            # Classify each flag against previous claims (fuzzy match on same URL)
            for f in flags:
                f.status = _classify_new_flag(f, job.previous_claims)

            job.flags.extend(flags)
            pages_analysed += 1
            job.pages_scanned = pages_analysed

            await emit("result", {
                "url": url,
                "title": title,
                "flags": [
                    {
                        "claim_text": f.claim_text,
                        "breach_reason": f.breach_reason,
                        "severity": f.severity,
                        "remediation": f.remediation,
                        "status": f.status,
                    } for f in flags
                ],
                "running_total": len(job.flags),
                "new_total": sum(1 for x in job.flags if x.status == "NEW"),
                "known_total": sum(1 for x in job.flags if x.status == "KNOWN"),
                "pages_done": pages_analysed,
            })

            # Harvest more links
            for link in _extract_links(r.text, url):
                if link in seen:
                    continue
                if not _is_same_site(link, start_netloc, req.include_subdomains):
                    continue
                if not _looks_like_html(link):
                    continue
                seen.add(link)
                queue.append(link)

            if req.delay > 0:
                await asyncio.sleep(req.delay)

        # After the crawl, any previous claim that wasn't matched is RESOLVED
        resolved_count = 0
        for prev in job.previous_claims:
            if prev.matched:
                continue
            job.flags.append(FlaggedClaim(
                url="https://" + prev.url_key if prev.url_key else "",
                page_title=prev.page_title,
                claim_text=prev.claim_text,
                breach_reason=prev.breach_reason,
                severity=prev.severity or "Low",
                remediation=prev.remediation,
                status="RESOLVED",
            ))
            resolved_count += 1

        await emit("status", {"message": "Building Excel report..."})
        job.excel_bytes = build_excel(job.flags, job.pages_scanned, req.url)

        job.status = "done"
        job.finished_at = datetime.utcnow()
        await emit("done", {
            "pages_scanned": job.pages_scanned,
            "total_flags": len(job.flags),
            "new_total": sum(1 for f in job.flags if f.status == "NEW"),
            "known_total": sum(1 for f in job.flags if f.status == "KNOWN"),
            "resolved_total": resolved_count,
            "download_url": f"/api/download/{job.id}",
        })


# --------------------------------------------------------------------------- #
# FastAPI app
# --------------------------------------------------------------------------- #

app = FastAPI(title="Greenwashing Auditor")


@app.exception_handler(RequestValidationError)
async def on_validation_error(request: Request, exc: RequestValidationError):
    # Print to stdout so the failing request body shows up in Railway logs
    try:
        body = (await request.body()).decode("utf-8", errors="replace")
    except Exception:
        body = "(could not read body)"
    print(f"[VALIDATION 422] {request.method} {request.url.path}")
    print(f"  body: {body[:500]}")
    print(f"  errors: {exc.errors()}")
    # Return a friendly, string-only detail the frontend can display
    return JSONResponse(
        status_code=422,
        content={"detail": exc.errors()},
    )

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
_static_dir = os.path.join(BASE_DIR, "static")
os.makedirs(_static_dir, exist_ok=True)
app.mount("/static", StaticFiles(directory=_static_dir), name="static")
templates = Jinja2Templates(directory=os.path.join(BASE_DIR, "templates"))


@app.get("/", response_class=HTMLResponse)
async def index(request: Request):
    return templates.TemplateResponse(request, "index.html")


# --------------------------------------------------------------------------- #
# Uploaded previous reports (short-lived, held only until start)
# --------------------------------------------------------------------------- #

# upload_id -> (claims, uploaded_at)
UPLOADS: dict[str, tuple[list[PreviousClaim], datetime]] = {}
UPLOAD_TTL_SECONDS = 30 * 60


def _sweep_old_uploads() -> None:
    cutoff = datetime.utcnow() - timedelta(seconds=UPLOAD_TTL_SECONDS)
    for uid, (_claims, ts) in list(UPLOADS.items()):
        if ts < cutoff:
            UPLOADS.pop(uid, None)


@app.get("/healthz")
async def healthz():
    return {"ok": True}


@app.post("/api/upload_previous")
async def upload_previous(request: Request):
    """Accept a multipart upload of a previous Excel report. Returns an
    upload_id that can be passed to /api/start to mark matching claims as KNOWN."""
    _sweep_old_uploads()
    form = await request.form()
    file = form.get("file")
    if file is None or not hasattr(file, "read"):
        raise HTTPException(status_code=400, detail="No file uploaded")

    filename = getattr(file, "filename", "") or ""
    if not filename.lower().endswith((".xlsx", ".xlsm")):
        raise HTTPException(status_code=400, detail="Please upload an .xlsx file")

    data = await file.read()
    if len(data) > 10 * 1024 * 1024:
        raise HTTPException(status_code=400, detail="File too large (10MB max)")

    try:
        claims = parse_previous_excel(data)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))

    upload_id = secrets.token_urlsafe(12)
    UPLOADS[upload_id] = (claims, datetime.utcnow())
    return {
        "upload_id": upload_id,
        "claim_count": len(claims),
        "filename": filename,
    }


@app.post("/api/start")
async def start_audit(req: StartAuditRequest):
    _sweep_old_jobs()

    # Basic URL sanity
    parsed = urlparse(req.url)
    if parsed.scheme not in ("http", "https") or not parsed.netloc:
        raise HTTPException(status_code=400, detail="Invalid URL - must include https://")

    # Clamp max_pages
    if req.max_pages > HARD_MAX_PAGES:
        req.max_pages = HARD_MAX_PAGES

    job_id = secrets.token_urlsafe(16)
    job = Job(id=job_id, request=req)

    # Attach previous claims if the user uploaded a prior report
    if req.previous_upload_id:
        entry = UPLOADS.pop(req.previous_upload_id, None)
        if entry is None:
            raise HTTPException(
                status_code=400,
                detail="Previous report upload has expired - please re-upload"
            )
        job.previous_claims = entry[0]

    JOBS[job_id] = job
    asyncio.create_task(run_audit(job))
    return {"job_id": job_id, "stream_url": f"/api/stream/{job_id}"}


@app.get("/api/stream/{job_id}")
async def stream(job_id: str, request: Request, since: int = 0):
    job = JOBS.get(job_id)
    if job is None:
        raise HTTPException(status_code=404, detail="Unknown job")

    async def event_gen() -> AsyncIterator[bytes]:
        yield b"retry: 2000\n\n"
        yield b": connected\n\n"

        # Replay any events the client missed before this connection opened.
        # The client tracks the highest `id:` it has seen and passes it as
        # ?since=N so we skip events 0..N-1 and resume from N.
        sent = 0
        replay = list(job.event_log)  # snapshot
        for idx, evt in enumerate(replay):
            if idx < since:
                continue
            payload = json.dumps(evt["data"])
            line = f"id: {idx}\nevent: {evt['event']}\ndata: {payload}\n\n".encode()
            yield line
            sent = idx + 1
            if evt["event"] in ("done", "error"):
                return  # job already finished, nothing more to send

        try:
            while True:
                if await request.is_disconnected():
                    break
                try:
                    evt = await asyncio.wait_for(job.queue.get(), timeout=5.0)
                except asyncio.TimeoutError:
                    yield b": heartbeat\n\n"
                    continue
                # Skip events already replayed (queue may still hold them)
                # by tracking the event_log length at emit time.
                current_idx = job.event_log.index(evt) if evt in job.event_log else sent
                if current_idx < sent:
                    continue
                payload = json.dumps(evt["data"])
                line = f"id: {current_idx}\nevent: {evt['event']}\ndata: {payload}\n\n".encode()
                yield line
                sent = current_idx + 1
                if evt["event"] in ("done", "error"):
                    break
        except asyncio.CancelledError:
            pass

    return StreamingResponse(
        event_gen(),
        media_type="text/event-stream",
        headers={
            "Cache-Control": "no-cache, no-transform",
            "X-Accel-Buffering": "no",
            "Connection": "keep-alive",
        },
    )


@app.get("/api/download/{job_id}")
async def download(job_id: str):
    job = JOBS.get(job_id)
    if job is None or job.excel_bytes is None:
        raise HTTPException(status_code=404, detail="Report not ready")

    safe_host = urlparse(job.request.url).netloc.replace(":", "_")
    filename = f"greenwashing_audit_{safe_host}_{job.id[:8]}.xlsx"
    return Response(
        content=job.excel_bytes,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
